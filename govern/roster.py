"""Roster intake: read + structurally validate the onboarding email/org file.

The onboarding roster is a CSV or Excel (.xlsx) file with a header row and one
or two columns:

  - **email** column (required) — the address to invite.
  - **group/org name** column (optional) — the organization to place them in.

Shape rules (enforced here; everything else — email/org *validity* — is checked
by the caller, which has the governed-org context):

  - More than 2 columns is a hard failure.
  - 1 column must be the email column (the org is chosen interactively later).
  - With 2 columns we auto-detect which is the email column and which is the org
    column (by which one actually contains valid email addresses).
  - The first row is treated as a header UNLESS it already looks like data (it
    contains a valid email or a valid org name) — in which case we warn and keep
    it as a data row.

Excel support imports ``openpyxl`` lazily so CSV-only users need no dependency.
"""
from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass, field
from typing import Callable, Optional

# Pragmatic email check: exactly one "@", non-empty local part, and a dotted
# domain with no spaces. Good enough to reject typos / stray org names without
# pretending to fully implement RFC 5322.
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class RosterError(Exception):
    """A hard, structural problem with the roster file (caller should exit)."""


def is_valid_email(value: Optional[str]) -> bool:
    return bool(value) and EMAIL_RE.match(value.strip()) is not None


@dataclass
class Roster:
    """A parsed roster: parallel ``emails`` / ``orgs`` lists (one per data row).

    ``orgs[i]`` is ``None`` for every row when the file had no org column (the
    caller resolves a single org interactively). ``warnings`` collects non-fatal
    notes (e.g. a missing header row) for the caller to surface.
    """

    emails: list[str]
    orgs: list[Optional[str]]
    has_org_column: bool
    warnings: list[str] = field(default_factory=list)

    def rows(self):
        return list(zip(self.emails, self.orgs))


def _strip_trailing_empty(row: list[str]) -> list[str]:
    out = list(row)
    while out and not out[-1].strip():
        out.pop()
    return out


def read_rows(path: str) -> list[list[str]]:
    """Read a .csv or .xlsx file into a list of string rows (cells stripped).

    Fully blank rows are dropped. Raises RosterError for unreadable / unsupported
    files so the caller can fail cleanly before doing anything else.
    """
    if not os.path.isfile(path):
        raise RosterError(f"roster file not found: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        rows = _read_csv(path, delimiter="\t" if ext == ".tsv" else ",")
    elif ext == ".xlsx":
        rows = _read_xlsx(path)
    elif ext in (".xls", ".xlsm", ".xlsb"):
        raise RosterError(
            f"unsupported Excel format {ext!r}; please save as .xlsx (or export to .csv)")
    else:
        raise RosterError(f"unsupported file type {ext!r}; use .csv or .xlsx")

    cleaned = []
    for row in rows:
        cells = [("" if c is None else str(c)).strip() for c in row]
        if any(cells):
            cleaned.append(cells)
    if not cleaned:
        raise RosterError(f"roster file is empty: {path}")
    return cleaned


def _read_csv(path: str, delimiter: str = ",") -> list[list[str]]:
    # utf-8-sig transparently strips a BOM if Excel wrote one.
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f, delimiter=delimiter)]


def _read_xlsx(path: str) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as e:  # pragma: no cover - depends on install
        raise RosterError(
            "reading .xlsx requires openpyxl (pip install -r requirements.txt); "
            "alternatively export the sheet to .csv") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _looks_like_data(row: list[str], is_valid_org: Callable[[str], bool]) -> bool:
    """A header row should be plain labels; if any cell is already a real email
    or a real org name, the row is data, not a header."""
    return any(is_valid_email(c) or is_valid_org(c) for c in row)


def _detect_email_column(data: list[list[str]]) -> int:
    """Return the index (0/1) of the column that holds the email addresses.

    We pick the column with strictly more valid emails. Ties (both columns look
    like emails, or neither does) are ambiguous and raise.
    """
    counts = [
        sum(1 for r in data if len(r) > col and is_valid_email(r[col]))
        for col in (0, 1)
    ]
    if counts[0] == 0 and counts[1] == 0:
        raise RosterError(
            "could not find a column of email addresses (neither column contains "
            "valid emails)")
    if counts[0] == counts[1]:
        raise RosterError(
            "could not auto-detect the email column (both columns look like "
            "emails); please provide email + group name columns")
    return 0 if counts[0] > counts[1] else 1


def parse_roster(path: str, *, is_valid_org: Callable[[str], bool]) -> Roster:
    """Read and structurally parse a roster file into a :class:`Roster`.

    Enforces the shape rules (<=2 columns, header detection, email-column
    detection). Email/org *value* validity is the caller's job — it owns the
    governed-org list and the consolidated error report.
    """
    rows = read_rows(path)
    width = max(len(_strip_trailing_empty(r)) for r in rows)
    if width > 2:
        raise RosterError(
            f"roster has {width} columns; expected at most 2 (email and group "
            "name). Remove the extra column(s) and try again.")
    if width == 0:  # pragma: no cover - read_rows already guards empties
        raise RosterError("roster has no data")

    warnings: list[str] = []
    if _looks_like_data(rows[0], is_valid_org):
        warnings.append(
            "first row looks like data (it contains a valid email or group "
            "name), not a header — treating it as a data row.")
        data = rows
    else:
        data = rows[1:]
    if not data:
        raise RosterError("roster has a header but no data rows")

    if width == 1:
        emails = [r[0].strip() if r else "" for r in data]
        return Roster(emails=emails, orgs=[None] * len(emails),
                      has_org_column=False, warnings=warnings)

    email_col = _detect_email_column(data)
    org_col = 1 - email_col

    def cell(r, i):
        return r[i].strip() if len(r) > i else ""

    emails = [cell(r, email_col) for r in data]
    orgs = [cell(r, org_col) for r in data]
    return Roster(emails=emails, orgs=orgs, has_org_column=True, warnings=warnings)
